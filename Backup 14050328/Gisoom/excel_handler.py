import openpyxl

COLUMN_MAPPING = {
    'عنوان اصلی': 'title',
    'تصویر': 'image_url',
    'نویسنده': 'author',
    'مترجم': 'translator',
    'ناشر': 'publisher',
    'صفحات': 'pages',
    'سال انتشار شمسی': 'year',
}


def get_header_map(ws):
    header_map = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            header_map[str(cell.value).strip()] = col_idx
    return header_map


def find_isbn_column(header_map):
    for name in ('شابک', 'ISBN', 'isbn'):
        if name in header_map:
            return header_map[name]
    return None


def update_excel(file_path, row_index, book_data):
    print("\n" + "=" * 50)
    print(f"[update_excel] فراخوانی شد | ردیف = {row_index}")
    print(f"[update_excel] کلیدهای دریافتی book_data = {list(book_data.keys())}")
    print(f"[update_excel] محتوای کامل book_data = {book_data}")

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"❌ باز کردن فایل شکست خورد: {e}")
        raise
    ws = wb.active

    header_map = get_header_map(ws)
    print(f"[update_excel] هدرهای پیداشده = {header_map}")

    written = 0
    for col_name, data_key in COLUMN_MAPPING.items():
        if col_name not in header_map:
            print(f"⚠️ ستون '{col_name}' در هدر اکسل نیست → رد شد")
            continue

        value = book_data.get(data_key)
        col_idx = header_map[col_name]

        if value:
            ws.cell(row=row_index, column=col_idx).value = value
            written += 1
            print(f"   ✓ نوشتم → ستون '{col_name}' (col={col_idx}) = {value!r}")
        else:
            print(f"   ✗ خالی/None برای کلید '{data_key}' "
                  f"(ستون '{col_name}') → نوشته نشد")

    try:
        wb.save(file_path)
        print(f"💾 ذخیره موفق | {written} فیلد در ردیف {row_index} نوشته شد")
    except PermissionError:
        print("❌ ذخیره شکست خورد: فایل اکسل احتمالاً همین الان در Excel باز است!")
        raise
    except Exception as e:
        print(f"❌ ذخیره شکست خورد: {e}")
        raise

    print("=" * 50)
    return written
