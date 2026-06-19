import openpyxl
from gisoom_crawler import GisoomCrawler
from excel_handler import update_excel


def normalize_isbn(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    # حذف کاراکترهای غیرضروری مثل فاصله، خط تیره، و برخی کاراکترهای مخفی
    s = s.replace(" ", "").replace("-", "").replace("\u202b", "").replace("\u202a", "")
    return s


def run_process(file_path):
    crawler = GisoomCrawler()

    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f"❌ فایل اکسل پیدا نشد: {file_path}")
        return
    except Exception as e:
        print(f"❌ خطا در باز کردن فایل اکسل: {e}")
        return

    ws = wb.active

    # پیدا کردن ستون‌ها از روی نام
    col_map = {cell.value: cell.column for cell in ws[1]}
    print(f"ℹ️ ستون‌های موجود: {list(col_map.keys())}")

    isbn_col = col_map.get("شابک")
    code_col = col_map.get("کد")

    if isbn_col is None:
        print("❌ ستون 'شابک' در ردیف اول پیدا نشد. لطفاً نام ستون را بررسی کنید.")
        return

    print(f"✅ ستون 'شابک' پیدا شد در شماره ستون: {isbn_col}")
    if code_col is not None:
        print(f"✅ ستون 'کد' پیدا شد در شماره ستون: {code_col}")

    total = 0
    success = 0
    not_found = 0
    invalid_isbn = 0

    for row in range(2, ws.max_row + 1):
        code = ws.cell(row=row, column=code_col).value if code_col else None
        raw_isbn = ws.cell(row=row, column=isbn_col).value
        isbn = normalize_isbn(raw_isbn)

        if not isbn:
            invalid_isbn += 1
            print(f"⚠️ ردیف {row} | کد={code} | شابک خالی یا نامعتبر است: {raw_isbn!r}")
            continue

        total += 1
        print(f"\n🔎 ردیف {row} | کد={code} | در حال جستجو برای ISBN: {isbn}")

        try:
            search_res = crawler.find_book_page(isbn)
        except Exception as e:
            print(f"❌ ردیف {row} | کد={code} | خطا در find_book_page برای ISBN={isbn}: {e}")
            continue

        if not search_res:
            not_found += 1
            print(f"❌ ردیف {row} | کد={code} | ISBN={isbn} | در گیسوم نتیجه‌ای پیدا نشد.")
            continue

        if not isinstance(search_res, dict) or "url" not in search_res:
            not_found += 1
            print(f"❌ ردیف {row} | کد={code} | ISBN={isbn} | خروجی find_book_page نامعتبر است: {search_res}")
            continue

        try:
            details = crawler.parse_book_page(search_res["url"])
        except Exception as e:
            print(f"❌ ردیف {row} | کد={code} | ISBN={isbn} | خطا در parse_book_page: {e}")
            continue

        if details is None:
            details = {}

        final_data = {**search_res, **details}

        try:
            update_excel(file_path, row, final_data)
            success += 1
            print(f"✅ ردیف {row} | کد={code} | ISBN={isbn} | اطلاعات ذخیره شد.")
        except Exception as e:
            print(f"❌ ردیف {row} | کد={code} | ISBN={isbn} | خطا در update_excel: {e}")

    

    print("\n==================== خلاصه ====================")
    print(f"کل ISBNهای پردازش‌شده: {total}")
    print(f"موفق: {success}")
    print(f"پیدا نشد: {not_found}")
    print(f"شابک نامعتبر/خالی: {invalid_isbn}")
    print("================================================")


if __name__ == "__main__":
    run_process(r"d:\python\Parsa Books\Gisoom\Parsa Library.xlsx")
